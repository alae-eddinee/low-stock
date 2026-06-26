import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from io import BytesIO

LIST_FILE = "list/ARTICLE PERMANENT GMS CADRE MIROIR MEUBLE.xlsx"
LOW_STOCK_THRESHOLD = 100

RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
RED_FONT   = Font(name="Calibri", size=11, color="FFFFFF")

BLUE_FILL  = PatternFill(start_color="DBFFC2", end_color="DBFFC2", fill_type="solid")  # commande header
LBLUE_FILL = PatternFill(start_color="C2FFC7", end_color="C2FFC7", fill_type="solid")  # labo name row
COL_FILL   = PatternFill(start_color="C2FFE6", end_color="C2FFE6", fill_type="solid")  # column headers
GRAY_FILL  = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")  # alternating rows (very light)
TOTAL_FILL = PatternFill(start_color="C2FFC7", end_color="C2FFC7", fill_type="solid")  # total row

DARK_FONT   = Font(name="Calibri", size=12, bold=True, color="1B4332")
WHITE_FONT  = DARK_FONT  # alias — these pastels need dark text
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="1B4332")
BODY_FONT   = Font(name="Calibri", size=11)
TOTAL_FONT  = Font(name="Calibri", size=11, bold=True, color="1B4332")

THIN   = Side(style="thin", color="A8E6B8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

# Excel epoch used by openpyxl when deserializing numeric cells with date format
_EXCEL_EPOCH = datetime(1899, 12, 31)


def _to_num(val):
    """
    openpyxl reads cells whose format is 'date' as datetime even when they hold
    a plain number.  Convert back to the original Excel serial integer/float.
    """
    if not isinstance(val, datetime):
        return val
    serial = (val - _EXCEL_EPOCH).days
    # Excel wrongly treats 1900 as a leap year; compensate for dates after Feb 28
    if val >= datetime(1900, 3, 1):
        serial += 1
    return serial


# ── Low Stock helpers ────────────────────────────────────────────────────────

def load_stock_lookup(file) -> dict:
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    headers  = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    code_col = next((i for i, h in enumerate(headers) if h and str(h).strip() == "Code"), None)
    qte_col  = next((i for i, h in enumerate(headers) if h and str(h).strip() == "Qte site6"), None)
    if code_col is None or qte_col is None:
        raise ValueError("Stock file must have 'Code' and 'Qte site6' columns.")
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[code_col]
        qty  = row[qte_col]
        if code is not None:
            lookup[str(code).strip()] = qty
    wb.close()
    return lookup


def build_stock_excel(list_path: str, stock_lookup: dict) -> tuple[BytesIO, int, int]:
    wb = openpyxl.load_workbook(list_path)
    ws = wb.active
    headers  = {cell.value: cell.column for cell in ws[1]}
    qte_col  = headers.get("Qte site6")
    code_col = headers.get("Code")
    if not qte_col or not code_col:
        raise ValueError("List file must have 'Code' and 'Qte site6' columns.")
    total = ws.max_row - 1
    low_count = 0
    for row in ws.iter_rows(min_row=2):
        code_cell = row[code_col - 1]
        qte_cell  = row[qte_col  - 1]
        code = str(code_cell.value).strip() if code_cell.value is not None else ""
        if code in stock_lookup:
            qte_cell.value = stock_lookup[code]
        qty = qte_cell.value
        if isinstance(qty, (int, float)) and qty < LOW_STOCK_THRESHOLD:
            qte_cell.fill = RED_FILL
            qte_cell.font = RED_FONT
            low_count += 1
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, total, low_count


# ── Labo analysis helpers ────────────────────────────────────────────────────

def load_labo_orders(file) -> dict:
    """
    Read labo Excel (file-like or path).
    Returns {ndoc: {labo, date, items: [(code, designation, qty, prix_ttc, montant_ttc)]}}
    """
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    orders = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ndoc        = row[3]
        tiers       = row[5]
        date        = row[1]
        code        = row[8]
        designation = row[9]
        qty         = _to_num(row[11])
        prix_ttc    = _to_num(row[17])
        montant_ttc = _to_num(row[20])
        if ndoc is None:
            continue
        if ndoc not in orders:
            orders[ndoc] = {"labo": tiers or "", "date": date, "items": []}
        orders[ndoc]["items"].append((str(code).strip() if code is not None else "", designation or "", qty, prix_ttc, montant_ttc))
    wb.close()
    return orders


def _cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:          c.font          = font
    if fill:          c.fill          = fill
    if alignment:     c.alignment     = alignment
    if border:        c.border        = border
    if number_format: c.number_format = number_format
    # Prevent openpyxl from misidentifying custom formats as date formats
    if isinstance(value, (int, float)):
        c.data_type = 'n'
    return c


def build_labo_excel(labo_file) -> BytesIO:
    """Build a styled Excel analysis file from labo data."""
    orders = load_labo_orders(labo_file)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analyse Labo"

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16

    current_row = 1

    for ndoc in sorted(orders.keys()):
        order = orders[ndoc]
        labo  = order["labo"]
        date  = order["date"]
        items = order["items"]

        date_str = date.strftime("%d/%m/%Y") if hasattr(date, "strftime") else str(date)

        # Row 1 — commande number + date (merged, dark blue)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=5)
        _cell(ws, current_row, 1,
              f"Commande N° {ndoc}   |   {date_str}",
              font=WHITE_FONT, fill=BLUE_FILL, alignment=CENTER)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Row 2 — labo name (merged, medium blue)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=5)
        _cell(ws, current_row, 1, labo,
              font=WHITE_FONT, fill=LBLUE_FILL, alignment=CENTER)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Row 3 — column headers
        for col_idx, label in enumerate(
            ["Code Article", "Désignation", "Qté", "Prix Unitaire TTC", "Montant TTC"], start=1
        ):
            _cell(ws, current_row, col_idx, label,
                  font=HEADER_FONT, fill=COL_FILL, alignment=CENTER, border=BORDER)
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Data rows
        total_qty = total_montant = 0
        for i, (code, designation, qty, prix_ttc, montant_ttc) in enumerate(items):
            row_fill = GRAY_FILL if i % 2 == 0 else None
            _cell(ws, current_row, 1, code,
                  font=BODY_FONT, fill=row_fill, alignment=CENTER, border=BORDER)
            _cell(ws, current_row, 2, designation,
                  font=BODY_FONT, fill=row_fill, alignment=LEFT,  border=BORDER)
            _cell(ws, current_row, 3, qty,
                  font=BODY_FONT, fill=row_fill, alignment=CENTER, border=BORDER,
                  number_format="#,##0")
            _cell(ws, current_row, 4, prix_ttc,
                  font=BODY_FONT, fill=row_fill, alignment=RIGHT, border=BORDER,
                  number_format="#,##0.00")
            _cell(ws, current_row, 5, montant_ttc,
                  font=BODY_FONT, fill=row_fill, alignment=RIGHT, border=BORDER,
                  number_format="#,##0.00")
            total_qty     += qty or 0
            total_montant += montant_ttc or 0
            current_row   += 1

        # Total row
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=2)
        _cell(ws, current_row, 1, "TOTAL",
              font=TOTAL_FONT, fill=TOTAL_FILL, alignment=CENTER, border=BORDER)
        _cell(ws, current_row, 2, None,
              font=TOTAL_FONT, fill=TOTAL_FILL, border=BORDER)
        _cell(ws, current_row, 3, int(total_qty),
              font=TOTAL_FONT, fill=TOTAL_FILL, alignment=CENTER, border=BORDER,
              number_format="#,##0")
        _cell(ws, current_row, 4, None,
              font=TOTAL_FONT, fill=TOTAL_FILL, border=BORDER)
        _cell(ws, current_row, 5, total_montant,
              font=TOTAL_FONT, fill=TOTAL_FILL, alignment=RIGHT, border=BORDER,
              number_format="#,##0.00")
        ws.row_dimensions[current_row].height = 18
        current_row += 3   # gap between orders

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Streamlit app ────────────────────────────────────────────────────────────

st.set_page_config(page_title="Low Stock Checker", layout="wide")
st.title("Low Stock Checker")

tab_stock, tab_labo = st.tabs(["Low Stock", "Analyse Labo"])

# ── Tab 1 ────────────────────────────────────────────────────────────────────
with tab_stock:
    st.markdown(
        "Upload the **stock file** — the app updates the article list with fresh "
        "quantities and highlights stock below **100** in red."
    )
    uploaded_stock = st.file_uploader("Upload stock.xlsx", type=["xlsx"])
    if uploaded_stock:
        try:
            stock_lookup = load_stock_lookup(uploaded_stock)
        except ValueError as e:
            st.error(str(e))
            st.stop()
        try:
            excel_buf, total, low_count = build_stock_excel(LIST_FILE, stock_lookup)
        except FileNotFoundError:
            st.error(f"Article list not found at: `{LIST_FILE}`")
            st.stop()
        except ValueError as e:
            st.error(str(e))
            st.stop()

        c1, c2 = st.columns(2)
        c1.metric("Articles in list", total)
        c2.metric(f"Low stock (< {LOW_STOCK_THRESHOLD})", low_count)
        st.download_button(
            label="Download updated list",
            data=excel_buf,
            file_name="low_stock_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ── Tab 2 ────────────────────────────────────────────────────────────────────
with tab_labo:
    st.header("Analyse Labo — Commandes")
    st.markdown(
        "Upload your labo data file to generate a styled Excel report — "
        "one section per commande, sorted by order number."
    )

    uploaded_labo = st.file_uploader("Upload labo data (.xlsx)", type=["xlsx"], key="labo")

    if uploaded_labo:
        try:
            labo_buf = build_labo_excel(uploaded_labo)
            st.success(f"Analyse prête — {uploaded_labo.name}")
            st.download_button(
                label="Télécharger l'analyse labo (Excel)",
                data=labo_buf,
                file_name="analyse_labo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(str(e))
